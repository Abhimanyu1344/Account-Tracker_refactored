"""
backend/output/report_builder.py — Styled 8-sheet reconciliation Excel report.
================================================================================
Public API (frozen signature):
    build_reconciliation_report(combined, summaries) → bytes

Dead commented-out code (165 lines) removed per R7.
================================================================================
"""

"""
report_builder.py
---------------------------------------------------------------------------
Generates the styled 8-sheet reconciliation Excel report.
Plugs directly into app_test.py with zero changes to other modules.

INTEGRATION — replace the download block (lines ~901-922) in app_test.py:
---------------------------------------------------------------------------
    from report_builder import build_reconciliation_report

    # inside: if st.session_state.recon_ran and st.session_state.recon_results:
    #   combined  = results.get("combined", {})
    #   summaries = results.get("summaries", [])

    excel_bytes = build_reconciliation_report(combined=combined, summaries=summaries)

    st.download_button(
        label     = "📥 Export Reconciliation Report",
        data      = excel_bytes,
        file_name = "Reconciliation_Report.xlsx",
        mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key       = "dl_recon_styled",
    )
---------------------------------------------------------------------------
"""

import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.output.presentation import (
    INR,
    apply_statement_presentation,
    apply_debtors_presentation,
    apply_kb_gaps_presentation,
    apply_unresolvable_presentation,
    apply_suspense_presentation,
    translate_aging_buckets,
    translate_confidence,
)

# ---------------------------------------------------------------------------
# Palette
# ---------------------------------------------------------------------------
DARK_BLUE    = "1F3864"
MID_BLUE     = "2E75B6"
LIGHT_BLUE   = "DEEAF1"
GREEN        = "70AD47"
LIGHT_GREEN  = "E2EFDA"
ORANGE       = "ED7D31"
LIGHT_ORANGE = "FCE4D6"
RED          = "C00000"
LIGHT_RED    = "FFCCCC"
LIGHT_YELLOW = "FFF2CC"
GREY         = "F2F2F2"
WHITE        = "FFFFFF"

# ---------------------------------------------------------------------------
# Primitives
# ---------------------------------------------------------------------------

def _border(color="DDDDDD"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _fmt_date(v):
    if isinstance(v, pd.Timestamp) and pd.notna(v):
        return v.strftime("%d-%b-%Y")
    s = str(v).strip()
    return s if s not in ("nan", "NaT", "None", "") else ""


def _safe_sum(summaries, key):
    try:
        return sum(float(s.get(key, 0) or 0) for s in summaries)
    except Exception:
        return 0.0


def _safe_count(summaries, key):
    try:
        return sum(int(s.get(key, 0) or 0) for s in summaries)
    except Exception:
        return 0


def _header(ws, row, labels, bg=DARK_BLUE, fg=WHITE):
    for col, text in enumerate(labels, 1):
        c = ws.cell(row=row, column=col, value=text)
        c.font      = Font(name="Arial", bold=True, color=fg, size=10)
        c.fill      = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _border("AAAAAA")


def _data(ws, row, vals, bg=WHITE, bold=False, amt_cols=None):
    fg = WHITE if bg in (DARK_BLUE, RED) else "000000"
    for col, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font      = Font(name="Arial", bold=bold, size=9, color=fg)
        c.fill      = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(vertical="center")
        c.border    = _border()
        if amt_cols and col in amt_cols and isinstance(v, (int, float)):
            c.number_format = "#,##0.00"
            c.alignment     = Alignment(horizontal="right", vertical="center")


def _col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def _title_block(ws, title, subtitle="", max_col=19):
    end_col = get_column_letter(max(max_col, 1))
    ws.merge_cells(f"A1:{end_col}1")
    c = ws["A1"]
    c.value     = "   " + title
    c.font      = Font(name="Arial", bold=True, size=14, color=WHITE)
    c.fill      = PatternFill("solid", start_color=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    if subtitle:
        ws.merge_cells(f"A2:{end_col}2")
        s = ws["A2"]
        s.value     = subtitle
        s.font      = Font(name="Arial", size=10, color="444444", italic=True)
        s.fill      = PatternFill("solid", start_color="EEF3F8")
        s.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 18
        return 3
    return 2


def _section_bar(ws, row, text, span=16, bg=MID_BLUE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(name="Arial", bold=True, size=11, color=WHITE)
    c.fill      = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22
    return row + 1


def _alt(i):
    return WHITE if i % 2 == 0 else GREY


def _status_bg(status):
    s = status.value.upper() if hasattr(status, 'value') else str(status).upper()
    if s in ("CLEARED", "FULLY_APPLIED"):               return LIGHT_GREEN
    if s in ("PARTIALLY_CLEARED", "PARTIALLY_APPLIED"): return LIGHT_YELLOW
    if s == "OPEN":                                     return LIGHT_RED
    return WHITE


def _conf_bg(val):
    try:
        v = float(val)
        if v >= 0.88: return LIGHT_GREEN
        if v >= 0.70: return LIGHT_YELLOW
        return LIGHT_RED
    except Exception:
        return WHITE


def _empty_df(v):
    return v if isinstance(v, pd.DataFrame) and not v.empty else pd.DataFrame()


# ---------------------------------------------------------------------------
# Sheet 1 – Summary Dashboard
# ---------------------------------------------------------------------------

def _build_summary(wb, summaries):
    ws = wb.create_sheet("Summary Dashboard")
    ws.sheet_view.showGridLines = False
    r = _title_block(ws, "RECONCILIATION SUMMARY DASHBOARD",
                     "Report Date: " + datetime.today().strftime("%d-%b-%Y"))

    total_bills   = _safe_count(summaries, "debtors_bills")
    bills_cleared = _safe_count(summaries, "bills_cleared")
    bills_partial = _safe_count(summaries, "bills_partial")
    bills_open    = _safe_count(summaries, "bills_open")
    amt_cleared   = _safe_sum(summaries,   "total_cleared")
    amt_out       = _safe_sum(summaries,   "total_outstanding")
    susp_applied  = _safe_sum(summaries,   "suspense_applied")
    susp_rem      = _safe_sum(summaries,   "suspense_remaining")
    kb_gaps       = _safe_count(summaries, "kb_gaps")
    unresolvable  = _safe_count(summaries, "unresolvable")
    clear_rate    = (bills_cleared / total_bills * 100) if total_bills else 0.0

    kpis = [
        ("TOTAL BILLS",            f"{total_bills:,}",                     DARK_BLUE, LIGHT_BLUE),
        ("CLEARED",                f"{bills_cleared:,}",                   GREEN,     LIGHT_GREEN),
        ("PARTIALLY CLEARED",      f"{bills_partial:,}",                   ORANGE,    LIGHT_ORANGE),
        ("OPEN",                   f"{bills_open:,}",                      RED,       LIGHT_RED),
        ("AMOUNT CLEARED",         f"{INR}{amt_cleared:,.0f}",             GREEN,     LIGHT_GREEN),
        ("OUTSTANDING",            f"{INR}{amt_out:,.0f}",                 RED,       LIGHT_RED),
        ("SUSPENSE APPLIED",       f"{INR}{susp_applied:,.0f}",            MID_BLUE,  LIGHT_BLUE),
        ("SUSPENSE CARRY FORWARD", f"{INR}{susp_rem:,.0f}",                ORANGE,    LIGHT_ORANGE),
    ]

    r = _section_bar(ws, r, "KEY PERFORMANCE INDICATORS")
    col = 1
    for label, value, txt, bg in kpis:
        ws.merge_cells(start_row=r,   start_column=col, end_row=r,   end_column=col + 1)
        ws.merge_cells(start_row=r+1, start_column=col, end_row=r+1, end_column=col + 1)
        lc = ws.cell(row=r,   column=col, value=label)
        vc = ws.cell(row=r+1, column=col, value=value)
        for cell, sz in [(lc, 8), (vc, 13)]:
            cell.font      = Font(name="Arial", bold=True, size=sz, color=txt)
            cell.fill      = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        col += 2
    ws.row_dimensions[r].height   = 28
    ws.row_dimensions[r+1].height = 28
    r += 3

    # Clearance-rate banner
    r = _section_bar(ws, r, "CLEARANCE RATE")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    rate_color = GREEN if clear_rate >= 70 else (ORANGE if clear_rate >= 40 else RED)
    rc = ws.cell(row=r, column=1, value=f"{clear_rate:.1f}%  bills cleared")
    rc.font      = Font(name="Arial", bold=True, size=16, color=rate_color)
    rc.fill      = PatternFill("solid", start_color=LIGHT_BLUE)
    rc.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=10)
    dc = ws.cell(row=r, column=6,
                 value=f"KB Gaps: {kb_gaps:,}   |   Unresolvable: {unresolvable:,}")
    dc.font      = Font(name="Arial", size=11, color="444444")
    dc.fill      = PatternFill("solid", start_color=GREY)
    dc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 28
    r += 2

    if not summaries:
        return

    r = _section_bar(ws, r, "PER-VERTICAL BREAKDOWN")
    hdrs = ["Vertical", "Total Bills", "Cleared", "Partial", "Open",
            f"Amt Cleared ({INR})", f"Outstanding ({INR})",
            f"Suspense Applied ({INR})", f"Carry Forward ({INR})",
            "KB Gaps", "Unresolvable"]
    _header(ws, r, hdrs)
    r += 1

    for i, s in enumerate(summaries):
        vals = [
            s.get("vertical", s.get("Vertical", "")),
            s.get("debtors_bills", 0),    s.get("bills_cleared", 0),
            s.get("bills_partial", 0),    s.get("bills_open", 0),
            s.get("total_cleared", 0),    s.get("total_outstanding", 0),
            s.get("suspense_applied", 0), s.get("suspense_remaining", 0),
            s.get("kb_gaps", 0),          s.get("unresolvable", 0),
        ]
        _data(ws, r, vals, bg=_alt(i), amt_cols={6, 7, 8, 9})
        for col in [6, 7, 8, 9]:
            ws.cell(row=r, column=col).number_format = "#,##0"
        r += 1

    totals = [
        "TOTAL",
        _safe_count(summaries, "debtors_bills"),  _safe_count(summaries, "bills_cleared"),
        _safe_count(summaries, "bills_partial"),   _safe_count(summaries, "bills_open"),
        _safe_sum(summaries,   "total_cleared"),   _safe_sum(summaries,   "total_outstanding"),
        _safe_sum(summaries,   "suspense_applied"), _safe_sum(summaries,  "suspense_remaining"),
        _safe_count(summaries, "kb_gaps"),          _safe_count(summaries, "unresolvable"),
    ]
    _data(ws, r, totals, bg=DARK_BLUE, bold=True, amt_cols={6, 7, 8, 9})
    for col in [6, 7, 8, 9]:
        ws.cell(row=r, column=col).number_format = "#,##0"
        ws.cell(row=r, column=col).font = Font(name="Arial", bold=True, color=WHITE, size=9)

    _col_widths(ws, [22, 12, 10, 10, 10, 20, 20, 22, 22, 10, 12])


# ---------------------------------------------------------------------------
# Sheet 2 – Reconciliation Statement
# ---------------------------------------------------------------------------

def _build_statement(wb, df):
    ws = wb.create_sheet("Reconciliation Statement")
    ws.sheet_view.showGridLines = False
    count = len(df) if not df.empty else 0
    r = _title_block(ws, "RECONCILIATION STATEMENT",
                     f"{count:,} knock-off events across all verticals")

    if not count:
        ws.cell(row=r, column=1, value="No knock-off events recorded.")
        return

    # Keep raw confidence scores for colour-coding before presentation apply
    raw_conf = df["Match_Confidence"].tolist() if "Match_Confidence" in df.columns else []

    out = apply_statement_presentation(df, for_excel=True)
    cols = list(out.columns)

    amt_display = [f"Bill Amt ({INR})", f"Amount Settled ({INR})", f"Balance After ({INR})"]
    amt_set     = {cols.index(c)+1 for c in amt_display if c in cols}
    conf_idx    = cols.index("Confidence") + 1 if "Confidence" in cols else None

    _header(ws, r, cols)
    r += 1

    for i, (_, row) in enumerate(out.iterrows()):
        vals = [_fmt_date(row[c]) if "Date" in c else row.get(c, "") for c in cols]
        _data(ws, r, vals, bg=_alt(i), amt_cols=amt_set)
        for col in amt_set:
            ws.cell(row=r, column=col).number_format = "#,##0.00"
        if conf_idx and i < len(raw_conf):
            ws.cell(row=r, column=conf_idx).fill = PatternFill(
                "solid", start_color=_conf_bg(raw_conf[i])
            )
        r += 1

    _col_widths(ws, [12, 32, 14, 12, 16, 12, 18, 18, 14, 16, 26, 38, 70])


# ---------------------------------------------------------------------------
# Sheet 3 – Updated Debtors
# ---------------------------------------------------------------------------

def _build_updated_debtors(wb, df):
    ws = wb.create_sheet("Updated Debtors")
    ws.sheet_view.showGridLines = False
    r = _title_block(ws, "UPDATED DEBTORS AGEING",
                     "All bills with reconciliation status applied")

    if df.empty:
        ws.cell(row=r, column=1, value="No debtors data.")
        return

    # Sort before presentation: Open first, then Partial, then Cleared
    priority = {"OPEN": 0, "PARTIALLY_CLEARED": 1, "CLEARED": 2}
    df_s = df.copy()
    if "Status" in df_s.columns:
        _get_val = lambda x: x.value.upper() if hasattr(x, 'value') else str(x).upper()
        df_s["_p"] = df_s["Status"].apply(_get_val).map(priority).fillna(3)
        sort_cols = ["_p"] + (["Pending Amount"] if "Pending Amount" in df_s.columns else [])
        asc       = [True] + [False] * (len(sort_cols) - 1)
        df_s = df_s.sort_values(sort_cols, ascending=asc).drop(columns=["_p"])

    # Raw status values for colour-coding before translation
    raw_status = df_s["Status"].tolist() if "Status" in df_s.columns else []

    out = apply_debtors_presentation(df_s, for_excel=True)
    cols = list(out.columns)

    amt_cols_names = [f"Original Amount ({INR})", f"Settled ({INR})", f"Outstanding ({INR})"]
    amt_set = {cols.index(c)+1 for c in amt_cols_names if c in cols}
    status_idx = cols.index("Status") + 1 if "Status" in cols else None

    _header(ws, r, cols)
    r += 1

    for i, (_, row) in enumerate(out.iterrows()):
        raw_s = raw_status[i] if i < len(raw_status) else ""
        bg = _status_bg(raw_s)
        vals = [_fmt_date(row[c]) if "Date" in c else row.get(c, "") for c in cols]
        _data(ws, r, vals, bg=bg, amt_cols=amt_set)
        for col in amt_set:
            ws.cell(row=r, column=col).number_format = "#,##0.00"
        r += 1

    _col_widths(ws, [12, 40, 18, 12, 16, 18, 18, 18, 20, 16, 16, 14])


# ---------------------------------------------------------------------------
# Sheet 4 – Updated Suspense
# ---------------------------------------------------------------------------

def _build_updated_suspense(wb, df):
    ws = wb.create_sheet("Updated Suspense")
    ws.sheet_view.showGridLines = False
    r = _title_block(ws, "UPDATED SUSPENSE LEDGER",
                     "All suspense credits with application status")

    if df.empty:
        ws.cell(row=r, column=1, value="No suspense data.")
        return

    raw_status = df["Status"].tolist() if "Status" in df.columns else []

    out = apply_suspense_presentation(df, for_excel=True)
    cols = list(out.columns)

    amt_names = [f"Total Received ({INR})", f"Applied to Invoice ({INR})", f"Carry Forward ({INR})"]
    amt_set   = {cols.index(c)+1 for c in amt_names if c in cols}

    _header(ws, r, cols)
    r += 1
    for i, (_, row) in enumerate(out.iterrows()):
        raw_s = raw_status[i] if i < len(raw_status) else ""
        bg    = _status_bg(raw_s) if "Status" in cols else _alt(i)
        vals  = [_fmt_date(row[c]) if c == "Bill Date" else row.get(c, "") for c in cols]
        _data(ws, r, vals, bg=bg, amt_cols=amt_set)
        for col in amt_set:
            ws.cell(row=r, column=col).number_format = "#,##0.00"
        r += 1

    _col_widths(ws, [12, 12, 28, 20, 20, 20, 20, 80])


# ---------------------------------------------------------------------------
# Sheet 5 – KB Gaps
# ---------------------------------------------------------------------------

def _build_kb_gaps(wb, df, df_debtors=None):
    ws = wb.create_sheet("KB Gaps")
    ws.sheet_view.showGridLines = False
    count = len(df) if not df.empty else 0
    r = _title_block(ws, "KNOWLEDGE BASE GAPS",
                     f"{count:,} entries partially matched  —  fill Confirmed Client and update KB")

    if not count:
        c = ws.cell(row=r, column=1, value="No KB gaps — all entries resolved.")
        c.font = Font(name="Arial", bold=True, color=GREEN, size=11)
        return

    out = apply_kb_gaps_presentation(
        df,
        updated_debtors_df=df_debtors if df_debtors is not None else pd.DataFrame(),
        for_excel=True,
    )
    cols = list(out.columns)
    conf_idx = cols.index("Confirmed_Client") + 1 if "Confirmed_Client" in cols else None
    amt_set  = {cols.index(f"Payment Amount ({INR})")+1} if f"Payment Amount ({INR})" in cols else set()

    _header(ws, r, cols, bg=MID_BLUE)
    r += 1
    for i, (_, row) in enumerate(out.iterrows()):
        vals = [_fmt_date(row[c]) if c == "Bill Date" else row.get(c, "") for c in cols]
        _data(ws, r, vals, bg=_alt(i), amt_cols=amt_set)
        for col in amt_set:
            ws.cell(row=r, column=col).number_format = "#,##0.00"
        if conf_idx:
            ws.cell(row=r, column=conf_idx).fill = PatternFill("solid", start_color=LIGHT_YELLOW)
        r += 1

    _col_widths(ws, [12, 12, 38, 20, 38, 36, 30, 18, 28, 80, 30])


# ---------------------------------------------------------------------------
# Sheet 6 – Unresolvable
# ---------------------------------------------------------------------------

def _build_unresolvable(wb, df):
    ws = wb.create_sheet("Unresolvable")
    ws.sheet_view.showGridLines = False
    count = len(df) if not df.empty else 0
    r = _title_block(ws, "UNRESOLVABLE ENTRIES",
                     f"{count:,} entries with no client candidate  —  manual review required")

    if not count:
        c = ws.cell(row=r, column=1, value="No unresolvable entries.")
        c.font = Font(name="Arial", bold=True, color=GREEN, size=11)
        return

    out  = apply_unresolvable_presentation(df, for_excel=True)
    cols = list(out.columns)
    amt_set = {cols.index(f"Amount ({INR})")+1} if f"Amount ({INR})" in cols else set()

    _header(ws, r, cols, bg=RED, fg=WHITE)
    r += 1
    for i, (_, row) in enumerate(out.iterrows()):
        vals = [_fmt_date(row[c]) if c == "Bill Date" else row.get(c, "") for c in cols]
        _data(ws, r, vals, bg=LIGHT_RED if i % 2 == 0 else WHITE, amt_cols=amt_set)
        for col in amt_set:
            ws.cell(row=r, column=col).number_format = "#,##0.00"
        r += 1

    _col_widths(ws, [12, 12, 18, 38, 36, 80, 18])


# ---------------------------------------------------------------------------
# Sheet 7 – Aging Analysis
# ---------------------------------------------------------------------------

def _build_aging(wb, df):
    ws = wb.create_sheet("Aging Analysis")
    ws.sheet_view.showGridLines = False
    r = _title_block(ws, "AGING ANALYSIS  —  REMAINING OUTSTANDING",
                     "Distribution of unpaid amounts across aging buckets by vertical")

    if df.empty or "Days Bucket" not in df.columns or "Vertical" not in df.columns:
        ws.cell(row=r, column=1, value="Aging bucket or Vertical data not available.")
        return

    amt_col = next((c for c in ["Remaining_Amount", "Pending Amount"] if c in df.columns), None)
    if not amt_col:
        ws.cell(row=r, column=1, value="No amount column found.")
        return

    df = df.copy()
    df[amt_col] = pd.to_numeric(df[amt_col], errors="coerce").fillna(0)

    pivot = (df.groupby(["Vertical", "Days Bucket"])[amt_col]
               .sum().unstack(fill_value=0).reset_index())

    # Put buckets in chronological order matching bucket_days() output format
    known_order = [
        "A: 0 - 30",
        "B: 31 - 45",
        "C: 46 - 60",
        "D: 61 - 90",
        "E: 91 - 120",
        "F: 121 - 150",
        "G: 151 - 180",
        "H: 181 - 360",
        "I: > 360",
    ]
    ordered = [b for b in known_order if b in pivot.columns]
    rest    = [b for b in pivot.columns if b not in ordered and b != "Vertical"]
    pivot   = pivot[["Vertical"] + ordered + rest]
    pivot[f"Grand Total"] = pivot.iloc[:, 1:].sum(axis=1)
    pivot = translate_aging_buckets(pivot)

    _header(ws, r, list(pivot.columns))
    r += 1
    for i, (_, row) in enumerate(pivot.iterrows()):
        vals = list(row.values)
        _data(ws, r, vals, bg=_alt(i), amt_cols=set(range(2, len(vals) + 1)))
        for col in range(2, len(vals) + 1):
            ws.cell(row=r, column=col).number_format = "#,##0"
        r += 1

    grand = ["GRAND TOTAL"] + [pivot.iloc[:, i].sum() for i in range(1, len(pivot.columns))]
    _data(ws, r, grand, bg=DARK_BLUE, bold=True, amt_cols=set(range(2, len(grand) + 1)))
    for col in range(2, len(grand) + 1):
        ws.cell(row=r, column=col).number_format = "#,##0"
        ws.cell(row=r, column=col).font = Font(name="Arial", bold=True, color=WHITE, size=9)

    _col_widths(ws, [20] + [16] * (len(pivot.columns) - 1))


# ---------------------------------------------------------------------------
# Sheet 8 – Clearance Summary
# ---------------------------------------------------------------------------

def _build_clearance(wb, df):
    ws = wb.create_sheet("Clearance Summary")
    ws.sheet_view.showGridLines = False
    r = _title_block(ws, "CLEARANCE SUMMARY BY VERTICAL",
                     "Bill status breakdown and clearance rates per vertical")

    if df.empty or "Vertical" not in df.columns:
        ws.cell(row=r, column=1, value="No debtors data available.")
        return

    status_col  = "Status"          if "Status"          in df.columns else None
    amt_col     = ("Pending Amount" if "Pending Amount"  in df.columns else
                   "Remaining_Amount" if "Remaining_Amount" in df.columns else None)
    cleared_col = "Cleared_Amount"  if "Cleared_Amount"  in df.columns else None

    rows_out = []
    for vertical, grp in df.groupby("Vertical"):
        total = len(grp)
        # Handle both enum values and legacy strings
        _get_val = lambda x: x.value.upper() if hasattr(x, 'value') else str(x).upper()
        if status_col:
            status_vals = grp[status_col].apply(_get_val)
            cl = int((status_vals == "CLEARED").sum())
            pa = int((status_vals == "PARTIALLY_CLEARED").sum())
            op = int((status_vals == "OPEN").sum())
        else:
            cl = pa = op = 0
        rate   = (cl / total * 100) if total else 0.0
        out    = pd.to_numeric(grp[amt_col],     errors="coerce").fillna(0).sum() if amt_col else 0
        cl_amt = pd.to_numeric(grp[cleared_col], errors="coerce").fillna(0).sum() if cleared_col else 0
        rows_out.append({
            "Vertical": vertical, "Total Bills": total,
            "Cleared": cl, "Partially Cleared": pa, "Open": op,
            "Clearance Rate %": round(rate, 1),
            f"Cleared Amt ({INR})": cl_amt, f"Outstanding ({INR})": out,
        })

    if not rows_out:
        ws.cell(row=r, column=1, value="No data to display.")
        return

    _header(ws, r, list(rows_out[0].keys()))
    r += 1
    for i, g in enumerate(rows_out):
        rate = g["Clearance Rate %"]
        bg   = LIGHT_GREEN if rate >= 70 else (LIGHT_YELLOW if rate >= 40 else LIGHT_RED)
        vals = list(g.values())
        _data(ws, r, vals, bg=bg, amt_cols={7, 8})
        ws.cell(row=r, column=6).number_format = '0.0"%"'
        for col in [7, 8]:
            ws.cell(row=r, column=col).number_format = "#,##0"
        r += 1

    _col_widths(ws, [20, 12, 10, 18, 10, 16, 20, 20])


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build_reconciliation_report(combined: dict, summaries: list) -> bytes:
    """
    Build the styled 8-sheet reconciliation Excel report.

    Parameters
    ----------
    combined  : dict  —  run_all_verticals()["combined"]
                         keys: statement, updated_debtors, updated_suspense,
                               kb_gaps, unresolvable
    summaries : list  —  run_all_verticals()["summaries"]

    Returns
    -------
    bytes  —  ready for st.download_button(data=...)
    """
    combined  = combined  or {}
    summaries = summaries or []

    df_stmt     = _empty_df(combined.get("statement"))
    df_debtors  = _empty_df(combined.get("updated_debtors"))
    df_suspense = _empty_df(combined.get("updated_suspense"))
    df_gaps     = _empty_df(combined.get("kb_gaps"))
    df_unres    = _empty_df(combined.get("unresolvable"))

    wb = Workbook()
    wb.remove(wb.active)

    _build_summary(wb, summaries)
    _build_statement(wb, df_stmt)
    _build_updated_debtors(wb, df_debtors)
    _build_updated_suspense(wb, df_suspense)
    _build_kb_gaps(wb, df_gaps, df_debtors=df_debtors)
    _build_unresolvable(wb, df_unres)
    _build_aging(wb, df_debtors)
    _build_clearance(wb, df_debtors)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()