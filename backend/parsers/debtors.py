"""
backend/parsers/debtors.py — Debtors ageing file parser + enrichment.
================================================================================
Responsibility:
  - Parse Tally debtors ageing Excel exports into a standardised DataFrame.
  - Enrich with: party name standardisation (PNM), state mapping, bucketing.
  - Smart column detection (handles varying file layouts across verticals).

Does NOT handle: vertical detection, receipts/suspense parsing, reconciliation.

PNM, state mapping, and bucketing live here (not in separate modules) because
parse_debtor_file is their sole consumer. If a second consumer appears, extract.
================================================================================
"""

import os
import re
import json
import pandas as pd
from datetime import datetime, date
from pathlib import Path

from backend.config import DATA_DIR
from backend.common import _parse_date, _cell_to_str, _clean_tally_text


# ══════════════════════════════════════════════════════════════════════════════
#  SMART COLUMN DETECTOR
# ══════════════════════════════════════════════════════════════════════════════

SKIP_PARTY_NAMES = {"", "TOTAL", "GRAND TOTAL"}

_FIELD_PATTERNS = {
    "date":   ["date"],
    "ref_no": ["ref. no", "ref no", "inv. no", "inv no", "invoice no"],
    "party":  ["party's name", "party name", "party"],
    "amount": ["pending amt", "pending amount"],
}

_SPLIT_HEADER_TOP    = "pending"
_SPLIT_HEADER_BOTTOM = "amount"

_REQUIRED_FIELDS = {"date", "party", "amount"}


def _score_row(row, next_row=None):
    """Score a single row against _FIELD_PATTERNS.
    Returns col_map: {field: col_idx}
    """
    col_map = {}
    used_cols = set()

    for col_idx, cell in enumerate(row):
        cell_str = _cell_to_str(cell)

        if next_row and col_idx < len(next_row):
            below_str = _cell_to_str(next_row[col_idx])
        else:
            below_str = ""

        for field, patterns in _FIELD_PATTERNS.items():
            if field in col_map:
                continue

            matched = False

            if field == "amount":
                if any(pat in cell_str for pat in patterns):
                    matched = True
                elif (cell_str == _SPLIT_HEADER_TOP and
                      below_str == _SPLIT_HEADER_BOTTOM):
                    matched = True
            else:
                if any(pat in cell_str for pat in patterns):
                    matched = True

            if matched:
                if col_idx not in used_cols:
                    col_map[field] = col_idx
                    used_cols.add(col_idx)
                break

    return col_map


def _detect_columns(ws):
    """Scan the first 20 rows to find header row and column positions.
    Returns (data_start_row_1based, col_map) or (None, {}).
    """
    rows = list(ws.iter_rows(min_row=1, max_row=20, values_only=True))

    for i, row in enumerate(rows):
        next_row = rows[i + 1] if i + 1 < len(rows) else None
        col_map = _score_row(row, next_row)

        if not _REQUIRED_FIELDS.issubset(col_map.keys()):
            continue

        mapped_cols = list(col_map.values())
        if len(mapped_cols) != len(set(mapped_cols)):
            continue

        date_col = col_map["date"]
        data_start = None
        for j in range(i + 1, len(rows)):
            cell = rows[j][date_col] if date_col < len(rows[j]) else None
            if _parse_date(cell) is not None:
                data_start = j + 1
                break

        if data_start is None:
            data_start = i + 2

        return data_start, col_map

    return None, {}


# ══════════════════════════════════════════════════════════════════════════════
#  JSON REFERENCE DATA LOADERS
# ══════════════════════════════════════════════════════════════════════════════

def _load_json(filename):
    path = DATA_DIR / filename
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"⚠  Reference file not found: {path} — using empty fallback")
        return {}
    except Exception as e:
        print(f"⚠  Failed to load {path}: {e} — using empty fallback")
        return {}


PARTY_NAME_MAP = _load_json('party_name_map.json')
STATE_MAP      = _load_json('state_mapping.json')


# ══════════════════════════════════════════════════════════════════════════════
#  PARTY NAME STANDARDISATION
# ══════════════════════════════════════════════════════════════════════════════

_PNM_NORMALISED = {k.strip().upper(): v for k, v in PARTY_NAME_MAP.items()}


def standardise_party_name(name):
    """Standardise party name using PNM lookup.
    Tries exact match first, then case-insensitive fallback.
    """
    if not name:
        return name
    cleaned = str(name).strip()
    exact = PARTY_NAME_MAP.get(cleaned)
    if exact:
        return exact
    normalised = _PNM_NORMALISED.get(cleaned.upper())
    if normalised:
        return normalised
    return cleaned


def extract_party_prefix(name):
    """Extract the part of the party name before the first - or _ delimiter."""
    if not name:
        return name
    s = str(name).strip()
    match = re.search(r'[-_]', s)
    if match and match.start() > 0:
        prefix = s[:match.start()].strip()
        return prefix.strip(". ").strip()
    return s.strip(". ").strip()


def extract_ref_prefix(ref_no):
    """Extract the alphabetic prefix from a ref number.
    Examples: CPD/25-26/131 → CPD, CPV-25-26/100 → CPV
    """
    if not ref_no:
        return None
    s = str(ref_no).strip().upper()
    match = re.match(r'^([A-Z_]+?)(?=[/\-]?\d{2}[-/])', s)
    if match:
        return match.group(1).strip('_-/')
    match = re.match(r'^([A-Z_]+)', s)
    return match.group(1).strip('_-/') if match else None


# ══════════════════════════════════════════════════════════════════════════════
#  STATE MAPPING
# ══════════════════════════════════════════════════════════════════════════════

def _normalise_key(s):
    """Produce a deterministic join key from a party name."""
    s = str(s).strip()
    s = re.sub(r'\s*[_]\s*', '_', s)
    s = re.sub(r'\s*[-]\s*', '-', s)
    s = re.sub(r'\s+', ' ', s)
    return s.upper()


_STATE_JOIN_TABLE = {_normalise_key(k): v for k, v in STATE_MAP.items()}


def map_state(party_name):
    """Map party name to state. Returns 'Not Found' if no match."""
    if not party_name:
        return "Not Found"
    key = _normalise_key(party_name)
    return _STATE_JOIN_TABLE.get(key, "Not Found")


# ══════════════════════════════════════════════════════════════════════════════
#  BUCKETING
# ══════════════════════════════════════════════════════════════════════════════

def bucket_days(days):
    """Map exact days overdue to a labelled slab."""
    if not isinstance(days, (int, float)):
        return ""
    d = int(days)
    if d <= 30: return "A: 0 - 30"
    if d <= 45: return "B: 31 - 45"
    if d <= 60: return "C: 46 - 60"
    if d <= 90: return "D: 61 - 90"
    if d <= 120: return "E: 91 - 120"
    if d <= 150: return "F: 121 - 150"
    if d <= 180: return "G: 151 - 180"
    if d <= 360: return "H: 181 - 360"
    return "I: > 360"


def bucket_amount(amount):
    """Map pending amount to a labelled slab."""
    if not isinstance(amount, (int, float)):
        return ""
    a = float(amount)
    if a <= 10000: return "A: 0 - 10K"
    if a <= 25000: return "B: 10.01 - 25K"
    if a <= 50000: return "C: 25.01 - 50K"
    if a <= 100000: return "D: 50.01 - 100K"
    if a <= 500000: return "E: 100.01 - 500K"
    return "F: > 500K"


# ══════════════════════════════════════════════════════════════════════════════
#  PARSER
# ══════════════════════════════════════════════════════════════════════════════

def parse_debtor_file(file_obj, vertical: str) -> pd.DataFrame:
    """Parse a Debtors Excel file and return a DataFrame with computed columns.

    Columns returned:
        Vertical, Date, Month, Ref. No., Party's Name,
        Pending Amount, Days Overdue, Days Bucket, Amount Bucket, State
    """
    import openpyxl

    file_obj.seek(0)
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)

    if "Bills Receivable" in wb.sheetnames:
        ws = wb["Bills Receivable"]
    else:
        ws = wb.active

    today = date.today()
    rows_data = []
    skipped_date_rows = []

    data_start, col_map = _detect_columns(ws)
    if not col_map:
        wb.close()
        return pd.DataFrame()

    date_idx   = col_map["date"]
    ref_idx    = col_map.get("ref_no")
    party_idx  = col_map["party"]
    amount_idx = col_map["amount"]

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        date_val = row[date_idx]
        ref_no   = row[ref_idx]  if ref_idx  is not None else None
        party    = row[party_idx]
        amount   = row[amount_idx]

        if not date_val or not amount:
            continue
        party_str = str(party).strip().upper() if party else ""
        if party_str in SKIP_PARTY_NAMES:
            continue

        bill_date = _parse_date(date_val)
        if not bill_date:
            skipped_date_rows.append(str(date_val))
            continue

        days_overdue = (today - bill_date).days
        raw_party    = str(party).strip() if party else ""
        std_party    = standardise_party_name(raw_party)
        # Use standardised name for state mapping — raw_party may have branch suffix (e.g. _JAIPUR)
        # that prevents lookup. std_party has already been normalised through PNM.
        state        = map_state(std_party)

        rows_data.append({
            "Vertical":       vertical,
            "Date":           bill_date.strftime("%d-%b-%Y"),
            "Month":          bill_date.strftime("%m-%Y"),
            "Ref. No.":       ref_no,
            "Party's Name":   std_party,
            "Pending Amount": amount,
            "Days Overdue":   days_overdue,
            "Days Bucket":    bucket_days(days_overdue),
            "Amount Bucket":  bucket_amount(amount),
            "State":          state,
        })

    wb.close()
    file_obj.seek(0)
    if skipped_date_rows:
        print(
            f"⚠  Debtors [{vertical}]: {len(skipped_date_rows)} rows skipped — "
            f"unparseable date values: {skipped_date_rows[:5]}"
            f"{'  ...' if len(skipped_date_rows) > 5 else ''}"
        )
    df = pd.DataFrame(rows_data)
    if not df.empty and "Pending Amount" in df.columns:
        non_numeric = df["Pending Amount"].apply(
            lambda x: not isinstance(x, (int, float))
        ).sum()
        if non_numeric > 0:
            print(f"⚠  Debtors [{vertical}]: {non_numeric} rows had non-numeric Pending Amount values")
    return df
