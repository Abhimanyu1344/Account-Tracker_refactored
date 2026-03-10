"""
backend/parsers/bank_books.py — Receipt + Suspense parsers (Tally bank books).
================================================================================
Responsibility:
  - Parse Tally receipt register exports (multi-sheet) into DataFrames.
  - Parse Tally suspense ledger exports (single-sheet) into DataFrames.
  - Shared Tally alternating-row merge logic (merge_tally_rows generator).
  - Vertical detection from filename / Excel row 1.
  - Standardised file naming.

Does NOT handle: debtors parsing, reconciliation, KB matching, presentation.

Receipts and suspense are both Tally bank-book exports sharing the same
alternating-row format, header detection, column layout, and skip-particular
rules. They live together because they share more code with each other than
with any other module.
================================================================================
"""

import os
import re
import pandas as pd
from datetime import datetime, date

from backend.common import (
    _parse_date, _clean_tally_text, _is_skip_particular,
    _find_header_row, _get_col_indices,
)


# ══════════════════════════════════════════════════════════════════════════════
#  VERTICAL DETECTION
# ══════════════════════════════════════════════════════════════════════════════

# Primary: entity name found in row 1 of the Excel file
ENTITY_MAP = [
    ("GREENFINCH REAL ESTATE ENG. & CON. (P) LTD. (H.O.)", "GREECPL"),
    ("GREENFINCH REAL ESTATE ENG", "GREECPL"),
    ("GREENFINCH REAL ESTATE ENGINEERS & CONSULTANTS", "GREEC"),
    ("GREENFINCH REAL ESTATE ENGINEERS", "GREEC"),
    ("GREENFINCH GLOBAL CONSULTING", "GFGC"),
    ("GREENFINCH TECH PROCESS", "Credit"),
    ("GREENFINCH LEGAL SERVICES", "Legal"),
]

# Fallback: keyword in filename
VERTICAL_MAP = [
    ("GF LEGAL", "Legal"),
    ("GF TECH", "Credit"),
    ("GREEC PVT. LTD.", "GREECPL"),
    ("GREEC PVT LTD", "GREECPL"),
    ("GREEC", "GREEC"),
    ("GFGC", "GFGC"),
]


def detect_vertical_from_row1(file_obj):
    """Scan every cell in row 1 of an Excel file for a known entity name.
    Returns (vertical, 'row1') or (None, None).
    """
    try:
        import openpyxl
        file_obj.seek(0)
        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        ws = wb.active
        first_row = [
            str(cell.value).strip().upper() if cell.value else ""
            for cell in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        wb.close()
        file_obj.seek(0)
        for cell_val in first_row:
            for keyword, vertical in ENTITY_MAP:
                if keyword.upper() in cell_val:
                    return vertical, "row1"
    except Exception:
        file_obj.seek(0)
    return None, None


def detect_vertical_from_filename(filename):
    """Match filename against known keywords.
    Returns (vertical, 'filename') or (None, None).
    """
    name_upper = filename.upper()
    for keyword, vertical in VERTICAL_MAP:
        if keyword.upper() in name_upper:
            return vertical, "filename"
    return None, None


def detect_vertical(file_obj, filename):
    """Try filename first, fall back to row 1.
    Returns (vertical, source) where source is 'row1' or 'filename'.
    """
    vertical, source = detect_vertical_from_filename(filename)
    if vertical:
        return vertical, source
    return detect_vertical_from_row1(file_obj)


def generate_file_name(file_type, original_name, vertical=None):
    """Generate a standardised saved filename."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    _, ext = os.path.splitext(original_name)
    prefix = file_type.upper()
    if vertical:
        return f"{prefix}_{vertical.upper().replace(' ', '_')}_{timestamp}{ext}"
    return f"{prefix}_{timestamp}{ext}"


# ══════════════════════════════════════════════════════════════════════════════
#  SUSPENSE ROUTING DETECTION
# ══════════════════════════════════════════════════════════════════════════════

_SUSPENSE_PREFIXES = ("SUSPENSE", "SUSPANC")


def _is_suspense_particular(val) -> bool:
    """
    Return True if Particulars signals a suspense-routed entry.

    Catches:
    1. Explicit suspense ledger names (Tally operator variants)
    2. Bank OD account descriptors with numeric account suffix
    """
    if val is None:
        return False
    s = str(val).strip().upper()

    if s == "SUSPENSE":
        return True
    for prefix in _SUSPENSE_PREFIXES:
        if s.startswith(prefix):
            return True

    if re.search(r'OD\s+(A/C|ACCOUNT)\s*\(?\d{4,}\)?', s):
        return True

    return False


# ══════════════════════════════════════════════════════════════════════════════
#  SHARED TALLY ROW MERGER (R3)
# ══════════════════════════════════════════════════════════════════════════════

def merge_tally_rows(raw_rows, cols):
    """
    Generator: yields (field_dict, narration_string_or_None) tuples.

    Encapsulates the alternating-row merge pattern shared by receipt and
    suspense parsers. Row N = transaction, Row N+1 = narration in the
    Particulars column.

    Skips rows where:
      - date cell is blank or literally "Date"
      - particular is in the unified skip-set
      - date is not parseable

    Args:
        raw_rows: list of tuples from ws.iter_rows(values_only=True)
        cols: dict with keys DATE, PARTICULAR, VCH_TYPE, VCH_NO, DEBIT, CREDIT
    """
    def get(r, idx):
        return r[idx] if idx < len(r) else None

    i = 0
    while i < len(raw_rows):
        row = raw_rows[i]

        date_val   = get(row, cols["DATE"])
        particular = get(row, cols["PARTICULAR"])
        vch_type   = get(row, cols["VCH_TYPE"])
        vch_no     = get(row, cols["VCH_NO"])
        debit      = get(row, cols["DEBIT"])
        credit     = get(row, cols["CREDIT"])

        # Skip blank or repeated header rows
        if not date_val or str(date_val).strip() == "Date":
            i += 1
            continue

        if _is_skip_particular(particular):
            i += 1
            continue

        bill_date = _parse_date(date_val)
        if not bill_date:
            i += 1
            continue

        # Check next row for narration
        narration = None
        if i + 1 < len(raw_rows):
            next_row        = raw_rows[i + 1]
            next_date       = get(next_row, cols["DATE"])
            next_particular = get(next_row, cols["PARTICULAR"])
            if not next_date and next_particular and not _is_skip_particular(next_particular):
                narration = _clean_tally_text(str(next_particular))
                i += 2
            else:
                i += 1
        else:
            i += 1

        fields = {
            "date":       bill_date,
            "particular":  particular,
            "vch_type":   vch_type,
            "vch_no":     vch_no,
            "debit":      debit,
            "credit":     credit,
        }

        yield fields, narration


# ══════════════════════════════════════════════════════════════════════════════
#  RECEIPT PARSER
# ══════════════════════════════════════════════════════════════════════════════

def parse_receipt_file(file_obj, vertical):
    """Parse a Receipt Register Excel file and return a DataFrame.

    Loops through ALL sheets (each sheet is a party/ledger in Tally export).
    Auto-detects header row position and 7-col vs 9-col layout per sheet.

    Columns returned:
        Vertical, Sheet, Date, Month, Vch Type, Vch No., Particulars, Narration, Debit, Credit
    """
    import openpyxl

    file_obj.seek(0)
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)

    all_rows = []

    for ws in wb.worksheets:
        sheet_name = ws.title

        header_row_num, row_len = _find_header_row(ws)
        if header_row_num is None:
            continue

        cols = _get_col_indices(row_len)
        data_start = header_row_num + 1

        raw_rows = []
        for row in ws.iter_rows(min_row=data_start, values_only=True):
            raw_rows.append(row)

        for fields, narration in merge_tally_rows(raw_rows, cols):
            bill_date  = fields["date"]
            particular = fields["particular"]

            all_rows.append({
                "Vertical":    vertical,
                "Sheet":       sheet_name,
                "Date":        bill_date.strftime("%d-%b-%Y"),
                "Month":       bill_date.strftime("%m-%Y"),
                "Vch Type":    str(fields["vch_type"]).strip() if fields["vch_type"] else "",
                "Vch No.":     str(fields["vch_no"]).strip()   if fields["vch_no"]   else "",
                "Particulars": "Suspense" if _is_suspense_particular(particular)
                               else _clean_tally_text(str(particular)) if particular else "",
                "Narration":   narration or "",
                "Debit":       fields["debit"]  if isinstance(fields["debit"],  (int, float)) else None,
                "Credit":      fields["credit"] if isinstance(fields["credit"], (int, float)) else None,
            })

    wb.close()
    file_obj.seek(0)
    return pd.DataFrame(all_rows)


# ══════════════════════════════════════════════════════════════════════════════
#  SUSPENSE PARSER
# ══════════════════════════════════════════════════════════════════════════════

def parse_suspense_file(file_obj, vertical: str) -> pd.DataFrame:
    """Parse a Suspense Excel file and return a DataFrame.

    Handles the Tally alternating-row export format (single sheet).

    Columns returned:
        Vertical, Date, Month, Particulars, Narration, Vch Type, Vch No., Debit, Credit
    """
    import openpyxl

    file_obj.seek(0)
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    if len(wb.worksheets) > 1:
        print(f"⚠  Suspense file [{vertical}] has {len(wb.worksheets)} sheets — only first sheet processed")

    # Suspense uses fixed 7-col layout: A=0, C=2, D=3, E=4, F=5, G=6
    COL_DATE       = 0
    COL_PARTICULAR = 2
    COL_VCH_TYPE   = 3
    COL_VCH_NO     = 4
    COL_DEBIT      = 5
    COL_CREDIT     = 6

    cols = dict(DATE=COL_DATE, PARTICULAR=COL_PARTICULAR, VCH_TYPE=COL_VCH_TYPE,
                VCH_NO=COL_VCH_NO, DEBIT=COL_DEBIT, CREDIT=COL_CREDIT)

    header_row_num, _ = _find_header_row(ws)
    if header_row_num is not None:
        data_start = header_row_num + 1
    else:
        print(f"⚠  Suspense [{vertical}]: Could not auto-detect header row — falling back to row 11")
        data_start = 11

    raw_rows = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        raw_rows.append(row)

    wb.close()
    file_obj.seek(0)

    rows_data = []
    non_numeric_count = 0

    for fields, narration in merge_tally_rows(raw_rows, cols):
        particular = fields["particular"]
        credit     = fields["credit"]
        debit      = fields["debit"]
        bill_date  = fields["date"]

        if credit is not None and not isinstance(credit, (int, float)):
            non_numeric_count += 1
        if debit is not None and not isinstance(debit, (int, float)):
            non_numeric_count += 1

        rows_data.append({
            "Vertical":    vertical,
            "Date":        bill_date.strftime("%d-%b-%Y"),
            "Month":       bill_date.strftime("%m-%Y"),
            "Particulars": _clean_tally_text(str(particular)) if particular else "",
            "Narration":   _clean_tally_text(narration) if narration else "",
            "Vch Type":    str(fields["vch_type"]).strip() if fields["vch_type"] else "",
            "Vch No.":     str(fields["vch_no"]).strip()   if fields["vch_no"]   else "",
            "Debit":       debit if isinstance(debit, (int, float)) else None,
            "Credit":      credit if isinstance(credit, (int, float)) else None,
        })

    if non_numeric_count > 0:
        print(f"⚠  Suspense [{vertical}]: {non_numeric_count} cells had non-numeric Debit/Credit values (set to None)")

    return pd.DataFrame(rows_data)
